"""
Excel Report Generator - Enhanced with formatting
Auto-fit columns, wrap text, conditional formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any, Optional


def create_excel_report(results: List[Dict[str, Any]], output_path: str) -> bool:
    """
    Create the simplest possible Excel report - just data in rows and columns
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        if not results:
            print("⚠ No results to export")
            return False
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Simple headers
        ws['A1'] = "Case Name"
        ws['B1'] = "Status"
        ws['C1'] = "Root Cause"
        ws['D1'] = "Diagnosis"
        ws['E1'] = "Suggested Fix"
        ws['F1'] = "Confidence"
        ws['G1'] = "Evidence"
        
        # Write data rows (starting from row 2)
        for idx, result in enumerate(results, start=2):
            # Case Name
            ws[f'A{idx}'] = result.get('case_name', 'unknown')
            
            # Status
            status = result.get('status', 'unknown')
            ws[f'B{idx}'] = 'CORRECT' if status == 'correct' else 'BROKEN'
            
            # Root Cause - SHORT SUMMARY
            root_cause = result.get('root_cause', 'N/A')
            if isinstance(root_cause, list):
                root_cause = root_cause[0] if root_cause else 'N/A'
            
            # Create short root cause
            root_cause_str = str(root_cause)
            if len(root_cause_str) > 100:
                # Shorten long root causes
                if "failed to load" in root_cause_str:
                    # Extract just the filename
                    import re
                    filename_match = re.search(r'([a-zA-Z0-9_.-]+\.(css|js))', root_cause_str)
                    if filename_match:
                        filename = filename_match.group(1)
                        ws[f'C{idx}'] = f"{filename} failed to load"
                    else:
                        ws[f'C{idx}'] = root_cause_str[:100] + "..."
                elif "repeat" in root_cause_str:
                    # Extract repetition count
                    rep_match = re.search(r'(\d+)x', root_cause_str)
                    if rep_match:
                        count = rep_match.group(1)
                        ws[f'C{idx}'] = f"Page repeated {count}x (duplicate stitching)"
                    else:
                        ws[f'C{idx}'] = "Duplicate stitching detected"
                else:
                    ws[f'C{idx}'] = root_cause_str[:100] + "..."
            else:
                ws[f'C{idx}'] = root_cause_str
            
            # Diagnosis - DETAILED TECHNICAL EXPLANATION
            diagnosis = result.get('diagnosis', [])
            if isinstance(diagnosis, list):
                diagnosis_text = "; ".join(diagnosis) if diagnosis else "See root cause"
            else:
                diagnosis_text = str(diagnosis)
            
            # Make diagnosis different from root cause by adding technical details
            if diagnosis_text == root_cause_str or not diagnosis_text or diagnosis_text == "See root cause":
                # Generate better diagnosis from evidence
                if status == 'correct':
                    diagnosis_text = "All visual metrics passed. No rendering issues detected."
                elif "repeat" in root_cause_str:
                    diagnosis_text = "CV detected sequential duplication pattern. Likely infinite scroll bug or screenshot capture error."
                elif "failed to load" in root_cause_str:
                    diagnosis_text = f"Critical resource missing. Page has HTML structure but rendering failed. Check network errors and file paths."
                else:
                    diagnosis_text = str(diagnosis)
            
            ws[f'D{idx}'] = diagnosis_text
            
            # Suggested Fix - ACTIONABLE DEVELOPER STEPS
            fixes = result.get('suggested_fix', [])
            if isinstance(fixes, list):
                fix_text = "; ".join(fixes) if fixes else "None"
            else:
                fix_text = str(fixes)
            
            # Confidence score
            confidence = result.get('confidence', 0.0)
            ws[f'F{idx}'] = f"{confidence:.2f}"
            
            # Improve generic fixes with specific steps
            if not fix_text or fix_text == "None" or "No action needed" in fix_text:
                if status == 'correct':
                    fix_text = "✅ No action needed - page is working correctly"
                else:
                    fix_text = "❌ Issue detected - see root cause for details"
            elif "Check" in fix_text and len(fix_text) < 50:
                # Expand vague "Check X" suggestions
                if "CSS" in fix_text or "css" in root_cause_str:
                    fix_text = "1. Check browser DevTools Network tab for 404 errors on CSS files\n2. Verify CSS file exists and path is correct\n3. Check for CORS issues if loading from CDN\n4. Re-deploy missing CSS bundle"
                elif "event listener" in fix_text or "infinite scroll" in fix_text:
                    fix_text = "1. Debug infinite scroll event listeners\n2. Check for duplicate event bindings\n3. Verify scroll detection logic\n4. Re-capture page with proper scroll handling"
                elif "component" in fix_text:
                    fix_text = "1. Check browser console for JS errors\n2. Verify web component registration\n3. Test component hydration in isolation\n4. Check framework version compatibility"
            
            ws[f'E{idx}'] = fix_text
            
            # Evidence - handle dict or string
            evidence = result.get('evidence', '')
            if isinstance(evidence, dict):
                # Combine all evidence into a readable string
                evidence_parts = []
                if evidence.get('cv_global_findings'):
                    evidence_parts.append("Global CV: " + "; ".join(evidence['cv_global_findings']))
                if evidence.get('cv_regional_findings'):
                    evidence_parts.append("Regional CV: " + "; ".join(evidence['cv_regional_findings']))
                if evidence.get('llm_findings'):
                    evidence_parts.append("LLM: " + "; ".join(evidence['llm_findings']))
                evidence_str = " | ".join(evidence_parts) if evidence_parts else 'N/A'
            else:
                evidence_str = str(evidence) if evidence else 'N/A'
            
            ws[f'G{idx}'] = evidence_str
            
            # Apply conditional formatting based on status
            if status == 'correct':
                # Green background for correct
                ws[f'B{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws[f'B{idx}'].font = Font(color="006100", bold=True)
            else:
                # Red background for broken
                ws[f'B{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws[f'B{idx}'].font = Font(color="9C0006", bold=True)
        
        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-fit column widths and enable text wrapping
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # Enable text wrapping
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        
                        # Calculate max length for auto-fit
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set column width (max 80 characters to avoid super wide columns)
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set specific column widths for better readability
        ws.column_dimensions['A'].width = 20  # Case Name
        ws.column_dimensions['B'].width = 12  # Status
        ws.column_dimensions['C'].width = 40  # Root Cause
        ws.column_dimensions['D'].width = 60  # Diagnosis
        ws.column_dimensions['E'].width = 50  # Suggested Fix
        ws.column_dimensions['F'].width = 12  # Confidence
        ws.column_dimensions['G'].width = 80  # Evidence
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save
        wb.save(output_path)
        print(f"✅ Simple Excel report saved: {output_path}")
        return True
    
    except Exception as e:
        print(f"❌ Error creating Excel report: {e}")
        print("  Falling back to CSV format...")
        return create_csv_report(results, output_path.replace('.xlsx', '.csv'))


def create_csv_report(results: List[Dict[str, Any]], output_path: str) -> bool:
    """
    Create CSV report for easy import into other tools
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        if not results:
            print("⚠ No results to export")
            return False
            
        import csv
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow(["Case Name", "Status", "Root Cause", "Diagnosis", "Suggested Fix", "Confidence", "Evidence"])
            
            # Data rows
            for result in results:
                case_name = result.get('case_name', 'unknown')
                status = 'CORRECT' if result.get('status') == 'correct' else 'BROKEN'
                root_cause = str(result.get('root_cause', 'N/A'))
                diagnosis = str(result.get('diagnosis', 'N/A'))
                suggested_fix = str(result.get('suggested_fix', 'N/A'))
                confidence = result.get('confidence', 0.0)
                
                # Handle evidence dict or string
                evidence = result.get('evidence', '')
                if isinstance(evidence, dict):
                    evidence_parts = []
                    if evidence.get('cv_global_findings'):
                        evidence_parts.append("Global: " + "; ".join(evidence['cv_global_findings']))
                    if evidence.get('cv_regional_findings'):
                        evidence_parts.append("Regional: " + "; ".join(evidence['cv_regional_findings']))
                    if evidence.get('llm_findings'):
                        evidence_parts.append("LLM: " + "; ".join(evidence['llm_findings']))
                    evidence = " | ".join(evidence_parts) if evidence_parts else 'N/A'
                else:
                    evidence = str(evidence) if evidence else 'N/A'
                
                writer.writerow([case_name, status, root_cause, diagnosis, suggested_fix, f"{confidence:.2f}", evidence])
        
        print(f"✅ CSV report saved: {output_path}")
        return True
    
    except Exception as e:
        print(f"❌ Error creating CSV report: {e}")
        return False


def export_to_excel(results: List[Dict[str, Any]], output_path: str) -> bool:
    """
    Wrapper function for Excel export with automatic CSV fallback
    
    Returns:
        bool: True if successful, False otherwise
    """
    return create_excel_report(results, output_path)
